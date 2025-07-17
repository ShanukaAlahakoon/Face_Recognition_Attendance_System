from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from flask_mysqldb import MySQL  # Add this import
from db_connection import (
    check_user_login,
    get_total_students,
    get_total_units,
    get_total_lecturers,
    get_lecturers,
    get_students,
    get_lecture_rooms,
    get_courses,
    get_units,
    get_total_faculties,
    get_faculties,
    get_db_connection,
    add_faculty_to_db,
    add_course_to_db,
    add_unit_to_db,
    get_lecturer_by_email,
    get_lecturer_courses,
    get_units_by_course,
    get_students_by_course,
    add_venue,
    add_lecturer_to_db,
    get_courses_by_faculty,
    add_student_to_db,
    get_venues_by_course,
    get_students_by_course_unit,
    get_available_venues_by_course,
    get_students_for_course_unit,
    get_attendance_for_unit
)
from functools import wraps
import MySQLdb
from data import capture_student_image
import os
import sys
import subprocess  # Add this import
import shutil  # Add this import at the top if not already present
import pandas as pd
import io
import signal
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'  # Add this if not already present

# MySQL Configuration
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'new11'

mysql = MySQL(app)  # Initialize MySQL

# Add this decorator definition before your routes
def role_required(required_role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'role' not in session:
                flash('Please log in first', 'error')
                return redirect(url_for('login'))
            if session['role'] != required_role:
                flash('Unauthorized access', 'error')
                return redirect(url_for('login'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.route('/')
def home():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        role = request.form.get('role')      
        email = request.form.get('email')    
        password = request.form.get('password')  

        print(f"\nLogin attempt - Role: {role}, Email: {email}")

        user = check_user_login(email, password, role)

        if user:
            session['email'] = user['emailAddress']
            session['role'] = role
            session['user_id'] = user['id']
            print(f"Session created: {session}")
            
            if role == 'Admin':
                return redirect(url_for('dashboard'))
            elif role == 'Lecturer':
                return redirect(url_for('lec_dashboard'))
        else:
            flash('Invalid Email, Password, or Role', 'error')

    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    if 'email' not in session:
        return redirect(url_for('login'))  # Prevent access if not logged in

    # Fetch data from the database
    total_students = get_total_students()
    total_units = get_total_units()
    total_lecturers = get_total_lecturers()
    lecturers = get_lecturers()
    students = get_students()
    lecture_rooms = get_lecture_rooms()
    courses = get_courses()

    print("Courses:", courses)

    total_courses = len(courses)  # or however you get the count

    return render_template('admin/admin_dashboard.html', 
                           total_students=total_students, 
                           total_units=total_units, 
                           total_lecturers=total_lecturers, 
                           lecturers=lecturers, 
                           students=students, 
                           lecture_rooms=lecture_rooms, 
                           courses=courses,
                           total_courses=total_courses)

@app.route('/logout')
def logout():
    session.clear()  # Clear all session data
    return redirect(url_for('login'))  # Redirect to login page

@app.route('/manage_courses')
def manage_courses():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # Courses with faculty name and counts
    cursor.execute("""
        SELECT 
            c.id,
            c.name,
            c.facultyID,
            (SELECT COUNT(*) FROM tblunit u WHERE u.courseID = c.courseCode) AS total_units,
            (SELECT COUNT(*) FROM tblstudents s WHERE s.courseCode = c.courseCode) AS total_students,
            DATE_FORMAT(c.dateCreated, '%Y-%m-%d') AS dateCreated
        FROM tblcourse c
        ORDER BY c.dateCreated DESC
    """)
    courses = cursor.fetchall()
    
    # Units with course name and student count
    cursor.execute("""
        SELECT 
            u.unitCode,
            u.name,
            c.name AS course,
            (SELECT COUNT(*) FROM tblstudents s WHERE s.courseCode = u.courseID) AS total_students,
            DATE_FORMAT(u.dateCreated, '%Y-%m-%d') AS dateCreated
        FROM tblunit u
        LEFT JOIN tblcourse c ON u.courseID = c.courseCode
        ORDER BY u.dateCreated DESC
    """)
    units = cursor.fetchall()
    
    # Faculties with course, student, and lecturer counts
    cursor.execute("""
        SELECT 
            f.facultyCode AS code,
            f.facultyName AS name,
            (SELECT COUNT(*) FROM tblcourse c WHERE c.facultyID = f.facultyCode) AS total_courses,
            (SELECT COUNT(*) FROM tblstudents s WHERE s.faculty = f.facultyCode) AS total_students,
            (SELECT COUNT(*) FROM tbllecturer l WHERE l.facultyCode = f.facultyCode) AS total_lectures,
            DATE_FORMAT(f.dateRegistered, '%Y-%m-%d') AS dateCreated
        FROM tblfaculty f
        ORDER BY f.dateRegistered DESC
    """)
    faculties = cursor.fetchall()
    
    cursor.close()
    return render_template('admin/course.html', courses=courses, units=units, faculties=faculties)

@app.route('/manage_lectures')
def manage_lectures():
    if 'email' not in session:
        return redirect(url_for('login'))

    # Fetch lecturers with their associated course
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("""
        SELECT
            l.id,
            l.firstName,
            l.lastName,
            l.emailAddress,
            l.phoneNo,
            l.facultyCode,
            c.name AS course  # Alias course name as 'course' to match template
        FROM tbllecturer l
        LEFT JOIN tblcourse c ON l.courseCode = c.courseCode # Join on courseCode
        ORDER BY l.firstName
    """)
    lecturers = cursor.fetchall()
    cursor.close()

    total_lecturers = len(lecturers) # Get total count from fetched data

    return render_template('admin/lecturer.html',
                         lecturers=lecturers,
                         total_lecturers=total_lecturers)

@app.route('/manage_students')
def manage_students():
    if 'email' not in session:
        return redirect(url_for('login'))
    
    students = get_students()
    total_students = get_total_students()
    
    return render_template('admin/student.html',
                         students=students,
                         total_students=total_students)

@app.route('/manage_venues')
def manage_venues():
    if 'email' not in session:
        return redirect(url_for('login'))
    
    venues = get_lecture_rooms()
    total_venues = len(venues)
    
    return render_template('admin/venue.html',
                         venues=venues,
                         total_venues=total_venues)

@app.route('/add_faculty', methods=['POST'])
def add_faculty():
    if 'email' not in session:
        return redirect(url_for('login'))
    
    faculty_name = request.form.get('facultyName')
    faculty_code = request.form.get('facultyCode')
    
    if add_faculty_to_db(faculty_code, faculty_name):
        flash('Faculty added successfully!', 'success')
    else:
        flash('Error adding faculty', 'error')
    
    return redirect(url_for('manage_courses'))

@app.route('/add_course', methods=['POST'])
def add_course():
    try:
        data = request.get_json()
        
        # Debug print
        print("Received data:", data)
        
        course_name = data.get('courseName')
        course_code = data.get('courseCode')
        faculty_id = data.get('faculty')
        
        # Validate input
        if not all([course_name, course_code, faculty_id]):
            return jsonify({
                'success': False, 
                'message': 'Missing required fields'
            })
        
        # Try to add the course
        result = add_course_to_db(course_name, course_code, faculty_id)
        
        if result:
            return jsonify({'success': True})
        else:
            return jsonify({
                'success': False, 
                'message': 'Failed to add course to database'
            })
            
    except Exception as e:
        print(f"Error in add_course route: {e}")
        return jsonify({
            'success': False, 
            'message': f'Error: {str(e)}'
        })

@app.route('/add_unit', methods=['POST'])
def add_unit():
    try:
        data = request.get_json()
        unit_code = data.get('unitCode')
        unit_name = data.get('unitName')
        course_id = data.get('courseId')
        
        print(f"Adding unit - Code: {unit_code}, Name: {unit_name}, Course: {course_id}")  # Debug print
        
        # Verify the course exists first
        cursor = mysql.connection.cursor()
        cursor.execute("SELECT courseCode FROM tblcourse WHERE courseCode = %s", (course_id,))
        course = cursor.fetchone()
        cursor.close()
        
        if not course:
            return jsonify({
                'success': False, 
                'message': f'Course with code {course_id} does not exist'
            })
        
        if add_unit_to_db(unit_code, unit_name, course_id):
            return jsonify({'success': True})
        else:
            return jsonify({
                'success': False, 
                'message': 'Failed to add unit to database'
            })
            
    except Exception as e:
        print(f"Error in add_unit route: {e}")  # Debug print
        return jsonify({
            'success': False, 
            'message': f'Error: {str(e)}'
        })

@app.route('/lec_dashboard')
def lec_dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
        
    email = session.get('email')
    lecturer = get_lecturer_by_email(email)
    courses = get_courses()
    faculties = get_faculties()  # <-- Add this line
    
    # Debug print
    print("Available courses:", courses)
    
    print("Lecturer object:", lecturer)
    if lecturer and 'venues' in lecturer:
        print("Lecturer venues:", lecturer['venues'])
    
    return render_template('lecturer/lec_dashboard.html', lecturer=lecturer, courses=courses, faculties=faculties)

@app.route('/get_units_by_course/<course_id>')
def fetch_units_by_course(course_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    print(f"Fetching units for course ID: {course_id}")  # Debug print
    units = get_units_by_course(course_id)
    print(f"Found units: {units}")  # Debug print
    return jsonify(units)

@app.route('/get_students_by_course_unit')
def fetch_students_by_course_unit():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    course_id = request.args.get('course')
    unit_id = request.args.get('unit')
    
    if not course_id or not unit_id:
        return jsonify([])
    
    print(f"Fetching students for course {course_id} and unit {unit_id}")  # Debug print
    students = sorted(get_students_by_course_unit(course_id, unit_id), key=lambda s: s['regNo'])
    print(f"Found students: {students}")  # Debug print
    
    return jsonify(students)

@app.route('/get_students_by_course/<course_id>')
def fetch_students_by_course(course_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    print(f"Fetching students for course ID: {course_id}")  # Debug print
    students = sorted(get_students_by_course(course_id), key=lambda s: s['regNo'])
    print(f"Found students: {students}")  # Debug print
    return jsonify(students)

@app.route('/get_available_venues_by_faculty/<faculty_code>')
def fetch_available_venues_by_faculty(faculty_code):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    print(f"Fetching available venues for faculty code: {faculty_code}")
    venues = get_available_venues_by_course(faculty_code)
    print(f"Found available venues: {venues}")
    return jsonify(venues)

@app.route('/add_venue', methods=['GET', 'POST'])
def add_venue_route():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        data = request.get_json()
        
        try:
            result = add_venue(
                className=data['className'],
                currentStatus=data['currentStatus'],
                capacity=data['capacity'],
                classType=data['classType'],
                faculty_code=data['facultyCode']
            )
            
            if result:
                return jsonify({'success': True})
            else:
                return jsonify({'success': False, 'message': 'Failed to add venue'})
                
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)})

@app.route('/get_faculties')
def get_faculties_route():
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute("""
            SELECT 
                facultyCode,
                facultyName
            FROM tblfaculty
            ORDER BY facultyName
        """)
        faculties = cursor.fetchall()
        cursor.close()
        
        print("Fetched faculties:", faculties)  # Debug print
        return jsonify(faculties)
        
    except Exception as e:
        print(f"Error fetching faculties: {e}")
        return jsonify([])

@app.route('/add_lecturer', methods=['POST'])
def add_lecturer():
    try:
        data = request.get_json()
        print("Received lecturer data:", data)  # Debug print

        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

        # Insert using the correct column names that match your table structure
        cursor.execute("""
            INSERT INTO tbllecturer (
                firstName,
                lastName,
                emailAddress,
                phoneNo,           # Changed from phoneNumber to phoneNo
                password,
                facultyCode,
                courseCode  # Added courseCode column
            ) VALUES (%s, %s, %s, %s, %s, %s, %s) # Added placeholder for courseCode
        """, (
            data['firstName'],
            data['lastName'],
            data['email'],
            data['phone'],
            data['password'],
            data['faculty'],
            data['course'] # Added course data
        ))

        mysql.connection.commit()
        cursor.close()

        return jsonify({'success': True})

    except Exception as e:
        print(f"Error adding lecturer: {str(e)}")  # Debug print
        return jsonify({
            'success': False,
            'message': f'Failed to add lecturer: {str(e)}'
        })

@app.route('/get_courses_by_faculty/<faculty_code>')
def get_courses_by_faculty_route(faculty_code):
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Updated query to get courses by faculty code
        cursor.execute("""
            SELECT 
                c.id,
                c.name,
                c.courseCode,
                f.facultyName
            FROM tblcourse c 
            LEFT JOIN tblfaculty f ON c.facultyID = f.facultyCode
            WHERE c.facultyID = %s
            ORDER BY c.name
        """, (faculty_code,))
        
        courses = cursor.fetchall()
        cursor.close()
        
        print(f"Fetched courses for faculty {faculty_code}:", courses)  # Debug print
        return jsonify(courses)
        
    except Exception as e:
        print(f"Error in get_courses_by_faculty_route: {e}")
        return jsonify([])

@app.route('/add_student', methods=['POST'])
def add_student():
    try:
        data = request.get_json()
        print("Received student data:", data)  # Debug print
        
        # Verify the course exists first
        connection = get_db_connection()
        cursor = connection.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT courseCode, facultyID 
            FROM tblcourse 
            WHERE courseCode = %s
        """, (data['courseId'],))
        course = cursor.fetchone()
        
        if not course:
            print(f"Course not found: {data['courseId']}")  # Debug print
            return jsonify({
                'success': False, 
                'message': f'Course with code {data["courseId"]} not found'
            })
        
        # Add the student
        result = add_student_to_db(
            first_name=data['firstName'],
            last_name=data['lastName'],
            email=data['email'],
            reg_no=data['registrationNumber'],
            course_id=data['courseId'],
            faculty=data['faculty']
        )
        
        if result:
            return jsonify({'success': True})
        else:
            return jsonify({
                'success': False, 
                'message': 'Failed to add student to database'
            })
            
    except Exception as e:
        print(f"Error in add_student route: {e}")  # Debug print
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals() and connection.is_connected():
            connection.close()

@app.route('/get_overview_counts')
def get_overview_counts():
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Get real-time counts
        cursor.execute("""
            SELECT 
                (SELECT COUNT(*) FROM tblcourse) as total_courses,
                (SELECT COUNT(*) FROM tblunit) as total_units,
                (SELECT COUNT(*) FROM tblfaculty) as total_faculties
        """)
        
        counts = cursor.fetchone()
        cursor.close()
        
        # Since we're using DictCursor, we can access by column names
        return jsonify({
            'total_courses': counts['total_courses'],
            'total_units': counts['total_units'],
            'total_faculties': counts['total_faculties']
        })
    except Exception as e:
        print(f"Error getting overview counts: {e}")
        return jsonify({
            'total_courses': 0,
            'total_units': 0,
            'total_faculties': 0
        })

@app.route('/get_courses_list')
def get_courses_list():
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Query to get all courses with their codes
        cursor.execute("""
            SELECT 
                id,
                name,
                courseCode,
                facultyID
            FROM tblcourse
            ORDER BY name
        """)
        
        courses = cursor.fetchall()
        cursor.close()
        
        print("Fetched courses for dropdown:", courses)
        return jsonify(courses)
        
    except Exception as e:
        print(f"Error fetching courses: {e}")
        return jsonify([])

@app.route('/get_courses_for_student')
def get_courses_for_student():
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Updated query to get all necessary course information
        cursor.execute("""
            SELECT 
                c.id,
                c.name,
                c.courseCode,
                f.facultyName,
                f.facultyCode
            FROM tblcourse c 
            LEFT JOIN tblfaculty f ON c.facultyID = f.facultyCode
            WHERE c.courseCode IS NOT NULL  # Ensure courseCode exists
            ORDER BY c.name
        """)
        
        courses = cursor.fetchall()
        cursor.close()
        
        print("Fetched courses for student form:", courses)  # Debug print
        return jsonify(courses)
        
    except Exception as e:
        print(f"Error fetching courses: {e}")  # Debug print
        return jsonify([])


@app.route('/get_venues_by_course/<int:course_id>')
def get_venues_by_course_route(course_id):
    try:
        venues = get_venues_by_course(course_id)
        return jsonify(venues)
    except Exception as e:
        print(f"Error in get_venues_by_course_route: {e}")
        return jsonify([])


@app.route('/capture_student_image', methods=['POST'])
def handle_capture_image():
    try:
        data = request.get_json()
        
        course = data.get('course')
        reg_number = data.get('reg_number')

        if not all([ course, reg_number]):
            return jsonify({
                'success': False,
                'message': 'Missing required fields'
            })

        success, message, image_paths = capture_student_image( course, reg_number)
        
        if success:
            # Convert the image paths to URL-friendly paths
            image_urls = []
            for path in image_paths:
                relative_path = os.path.relpath(path, 'static')
                image_urls.append(f'/static/{relative_path}')
            
            return jsonify({
                'success': True,
                'message': message,
                'image_paths': image_urls
            })
        else:
            return jsonify({
                'success': False,
                'message': message
            })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        })

attendance_process = None  # Global variable (for demo; use a better approach for production)

@app.route('/launch_attendance', methods=['POST'])
def launch_attendance():
    global attendance_process
    try:
        data = request.get_json()
        course_id = data.get('course_id')
        unit_id = data.get('unit_id')
        venue_id = data.get('venue_id')

        if not all([course_id, unit_id, venue_id]):
            return jsonify({
                'success': False,
                'error': 'Missing required parameters'
            })

        # Get course code from course ID and unit code from unit ID
        connection = get_db_connection()
        cursor = connection.cursor(dictionary=True)

        cursor.execute("SELECT courseCode FROM tblcourse WHERE id = %s", (course_id,))
        course = cursor.fetchone()

        cursor.execute("SELECT unitCode FROM tblunit WHERE id = %s", (unit_id,))
        unit = cursor.fetchone()

        cursor.close()
        connection.close()

        if not course:
            return jsonify({
                'success': False,
                'error': 'Course not found'
            })
        
        if not unit:
             return jsonify({
                 'success': False,
                 'error': 'Unit not found'
             })

        # Get the absolute path to Attendance.py
        attendance_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Attendance.py')
        try:
            process = subprocess.Popen(
                [sys.executable, attendance_script, course['courseCode'], unit['unitCode'], str(venue_id)],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1
            )
            attendance_process = process

            # Wait for "ENCODING_COMPLETE" in stdout
            while True:
                line = process.stdout.readline()
                if not line:
                    break
                print(line.strip())  # Optionally log output
                if "ENCODING_COMPLETE" in line:
                    break

        except Exception as e:
            print(f"Error launching Attendance.py: {e}")
            return jsonify({
                'success': False,
                'error': f'Failed to launch Attendance.py: {str(e)}'
            })

        # Only return success after encoding is complete
        return jsonify({
            'success': True,
            'message': f'Attendance system launched successfully for course {course["courseCode"]}, Unit {unit["unitCode"]}, Venue ID {venue_id}'
        })

    except Exception as e:
        print(f"Error launching attendance: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/get_attendance_status', methods=['GET'])
def get_attendance_status():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized access'}), 401

    course_id = request.args.get('course_id')
    unit_id = request.args.get('unit_id')
    venue_id = request.args.get('venue_id')

    if not all([course_id, unit_id, venue_id]):
        return jsonify({'error': 'Missing required parameters'}), 400

    try:
        connection = get_db_connection()
        cursor = connection.cursor(dictionary=True)

        # First, get the unitCode from the unit_id
        cursor.execute("SELECT unitCode FROM tblunit WHERE id = %s", (unit_id,))
        unit = cursor.fetchone()

        if not unit:
            cursor.close()
            connection.close()
            return jsonify({'error': 'Unit not found'}), 404

        unit_code = unit['unitCode']

        # Query to get attendance status for students in the selected course, unit, and venue for today
        # We need the student registration number and their attendance status
        query = """
        SELECT 
            s.registrationNumber, 
            CASE WHEN ta.attendanceStatus IS NOT NULL THEN ta.attendanceStatus ELSE 'Absent' END as attendanceStatus
        FROM tblstudents s
        JOIN tblcourse c ON s.courseCode = c.courseCode
        JOIN tblunit u ON u.courseID = c.courseCode
        LEFT JOIN tblattendance ta ON s.registrationNumber = ta.studentRegistrationNumber
                                    AND ta.unit = %s
                                    AND ta.venueId = %s
                                    AND DATE(ta.dateMarked) = CURDATE()
        WHERE c.id = %s AND u.id = %s
        """

        cursor.execute(query, (unit_code, venue_id, course_id, unit_id))
        attendance_data = cursor.fetchall()

        cursor.close()
        connection.close()

        # Convert list of dicts to a dictionary with registrationNumber as key for easy lookup in JS
        attendance_dict = {row['registrationNumber'].strip().upper(): row['attendanceStatus'] for row in attendance_data}

        return jsonify(attendance_dict)

    except Exception as e:
        print(f"Error fetching attendance status: {e}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/view_attendance')
def view_attendance():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('lecturer/view_attendance.html')

@app.route('/get_attendance_records')
def get_attendance_records():
    course_id = request.args.get('course_id')
    unit_id = request.args.get('unit_id')
    # 1. Get all students registered for the course/unit
    students = get_students_for_course_unit(course_id, unit_id)
    # 2. Get attendance records for the unit
    attendance = get_attendance_for_unit(unit_id)
    attendance_dict = {rec['registrationNumber']: rec for rec in attendance}
    # 3. Build the result
    result = []
    for student in students:
        rec = attendance_dict.get(student['registrationNumber'])
        if rec:
            result.append({
                'registrationNumber': student['registrationNumber'],
                'name': student['name'],
                'dateMarked': rec['dateMarked'],
                'attendanceStatus': rec['attendanceStatus']
            })
        else:
            result.append({
                'registrationNumber': student['registrationNumber'],
                'name': student['name'],
                'dateMarked': '',  # or the selected date
                'attendanceStatus': 'Absent'
            })
    return jsonify(result)

@app.route('/lec_student')
def lec_student():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('lecturer/lec_student.html')

@app.route('/download_attendance')
def download_attendance():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('lecturer/download_attendance.html')

@app.route('/get_attendance_matrix')
def get_attendance_matrix():
    course_id = request.args.get('course_id')
    unit_id = request.args.get('unit_id')
    if not course_id or not unit_id:
        return jsonify({"dates": [], "students": []})

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    # Get all unique dates for this course/unit
    cursor.execute("""
        SELECT DISTINCT DATE(ta.dateMarked) as dateMarked
        FROM tblattendance ta
        JOIN tblunit u ON ta.unit = u.unitCode
        JOIN tblcourse c ON u.courseID = c.courseCode
        WHERE c.id = %s AND u.id = %s
        ORDER BY dateMarked
    """, (course_id, unit_id))
    dates = [row['dateMarked'].strftime('%Y-%m-%d') for row in cursor.fetchall()]

    # Get all students for this course/unit, including names
    cursor.execute("""
        SELECT s.registrationNumber, CONCAT(s.firstName, ' ', s.lastName) as name
        FROM tblstudents s
        JOIN tblcourse c ON s.courseCode = c.courseCode
        JOIN tblunit u ON u.courseID = c.courseCode
        WHERE c.id = %s AND u.id = %s
    """, (course_id, unit_id))
    students = cursor.fetchall()

    # Get all attendance records for this course/unit
    cursor.execute("""
        SELECT ta.studentRegistrationNumber, DATE(ta.dateMarked) as dateMarked, ta.attendanceStatus
        FROM tblattendance ta
        JOIN tblunit u ON ta.unit = u.unitCode
        JOIN tblcourse c ON u.courseID = c.courseCode
        WHERE c.id = %s AND u.id = %s
    """, (course_id, unit_id))
    records = cursor.fetchall()
    cursor.close()
    connection.close()

    # Build matrix
    student_map = {reg['registrationNumber']: {"registrationNumber": reg['registrationNumber'], "name": reg['name'], "attendance": {}} for reg in students}
    for rec in records:
        reg = rec['studentRegistrationNumber']
        date = rec['dateMarked'].strftime('%Y-%m-%d')
        status = rec['attendanceStatus']
        if reg in student_map:
            student_map[reg]["attendance"][date] = status

    # Fill in "Absent" for missing dates
    for student in student_map.values():
        for date in dates:
            if date not in student["attendance"]:
                student["attendance"][date] = "Absent"

    return jsonify({
        "dates": dates,
        "students": list(student_map.values())
    })

@app.route('/delete_student', methods=['POST'])
def delete_student():
    reg_no = request.json.get('registrationNumber')
    if not reg_no:
        return jsonify({'success': False, 'message': 'Missing registration number'})
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        # Delete attendance records first
        cursor.execute("DELETE FROM tblattendance WHERE studentRegistrationNumber = %s", (reg_no,))
        # Now delete the student
        cursor.execute("DELETE FROM tblstudents WHERE registrationNumber = %s", (reg_no,))
        connection.commit()
        deleted = cursor.rowcount
        cursor.close()
        connection.close()
        # Delete the student's image folder
        image_folder = os.path.join('static', 'face_data', reg_no)
        try:
            if os.path.exists(image_folder):
                shutil.rmtree(image_folder)
        except Exception as e:
            print(f"Error deleting image folder for {reg_no}: {e}")
        if deleted > 0:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'message': 'Student not found or already deleted'})
    except Exception as e:
        print(f"Error deleting student: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/delete_lecturer', methods=['POST'])
def delete_lecturer():
    lecturer_id = request.json.get('id')
    if not lecturer_id:
        return jsonify({'success': False, 'message': 'Missing lecturer id'})
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("DELETE FROM tbllecturer WHERE id = %s", (lecturer_id,))
    connection.commit()
    cursor.close()
    connection.close()
    return jsonify({'success': True})

@app.route('/delete_room', methods=['POST'])
def delete_room():
    room_id = request.json.get('id')
    if not room_id:
        return jsonify({'success': False, 'message': 'Missing room id'})
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("DELETE FROM tblvenue WHERE id = %s", (room_id,))
    connection.commit()
    cursor.close()
    connection.close()
    return jsonify({'success': True})

@app.route('/delete_course', methods=['POST'])
def delete_course():
    course_id = request.json.get('id')
    if not course_id:
        return jsonify({'success': False, 'message': 'Missing course id'})
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("DELETE FROM tblcourse WHERE id = %s", (course_id,))
    connection.commit()
    cursor.close()
    connection.close()
    return jsonify({'success': True})

@app.route('/delete_unit', methods=['POST'])
def delete_unit():
    try:
        unit_code = request.json.get('unitCode')
        if not unit_code:
            return jsonify({'success': False, 'message': 'Missing unit code'})
            
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # First check if unit exists and get its course
        cursor.execute("""
            SELECT u.unitCode, u.courseID 
            FROM tblunit u 
            WHERE u.unitCode = %s
        """, (unit_code,))
        unit = cursor.fetchone()
        
        if not unit:
            cursor.close()
            connection.close()
            return jsonify({'success': False, 'message': 'Unit not found'})
            
        # Get all students in this course
        cursor.execute("""
            SELECT s.registrationNumber 
            FROM tblstudents s 
            WHERE s.courseCode = %s
        """, (unit[1],))  # unit[1] is courseID
        students = cursor.fetchall()
        
        if students:
            # Delete attendance records for these students first
            for student in students:
                cursor.execute("""
                    DELETE FROM tblattendance 
                    WHERE studentRegistrationNumber = %s AND unit = %s
                """, (student[0], unit_code))
            
            connection.commit()
        
        # Now delete the unit
        cursor.execute("DELETE FROM tblunit WHERE unitCode = %s", (unit_code,))
        connection.commit()
        
        deleted = cursor.rowcount
        cursor.close()
        connection.close()
        
        if deleted > 0:
            return jsonify({'success': True, 'message': 'Unit deleted successfully'})
        else:
            return jsonify({'success': False, 'message': 'Failed to delete unit'})
            
    except MySQLdb.Error as e:
        print(f"MySQL Error: {e}")
        return jsonify({'success': False, 'message': f'Database error: {str(e)}'})
    except Exception as e:
        print(f"Error deleting unit: {e}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/delete_faculty', methods=['POST'])
def delete_faculty():
    code = request.json.get('code')
    if not code:
        return jsonify({'success': False, 'message': 'Missing faculty code'})
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("DELETE FROM tblfaculty WHERE facultyCode = %s", (code,))
    connection.commit()
    cursor.close()
    connection.close()
    return jsonify({'success': True})

@app.route('/export_attendance_excel')
def export_attendance_excel():
    course_id = request.args.get('course_id')
    unit_id = request.args.get('unit_id')
    date = request.args.get('date')  # <-- Get the date from the request

    # Fetch attendance data for the given course/unit/date
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    cursor.execute("""
        SELECT s.registrationNumber, CONCAT(s.firstName, ' ', s.lastName) as name, c.name as course, u.name as unit,
               v.className as venue,  -- <--- Add this line
               IFNULL(a.attendanceStatus, 'Absent') as attendanceStatus
        FROM tblstudents s
        JOIN tblcourse c ON s.courseCode = c.courseCode
        JOIN tblunit u ON u.courseID = c.courseCode
        LEFT JOIN tblattendance a 
            ON a.studentRegistrationNumber = s.registrationNumber 
            AND a.unit = u.unitCode 
            AND DATE(a.dateMarked) = %s
        LEFT JOIN tblvenue v ON a.venueId = v.id  -- <--- Add this line
        WHERE c.id = %s AND u.id = %s
        ORDER BY s.registrationNumber
    """, (date, course_id, unit_id))
    data = cursor.fetchall()
    cursor.close()
    connection.close()

    # Prepare Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # 1. University name at the top (centered and bold)
    university_name = "South Eastern University of Sri Lanka"
    ws.merge_cells('A1:C1')  # Adjust range if you have more columns
    ws['A1'] = university_name
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 2. (Optional) Add a blank row for spacing
    ws.append([])

    # 3. Continue with your course/unit info and table as before
    course_name = data[0]['course'] if data else ''
    unit_name = data[0]['unit'] if data else ''
    venue_name = data[0].get('venue', '') if data else ''

    ws['A3'] = f"Date:- {date}"
    ws['A3'].font = Font(bold=True)
    ws['A3'] = f"Course Name:- {course_name}"
    ws['C3'] = f"Unit:- {unit_name}"
    ws['C3'].font = Font(bold=True)
    ws['A5'] = "Registration No"
    ws['B5'] = "Name"
    ws['C5'] = "Status"
    for col in ['A', 'B', 'C']:
        ws[f"{col}5"].font = Font(bold=True)
        ws[f"{col}5"].alignment = Alignment(horizontal='center')

    # Attendance data
    row_num = 6
    for row in data:
        ws[f"A{row_num}"] = row['registrationNumber']
        ws[f"B{row_num}"] = row['name']
        ws[f"C{row_num}"] = row['attendanceStatus']
        row_num += 1

    # Add a few empty rows for manual signature
    for _ in range(5):
        ws.append(['', '', ''])

    # Date and Signature rows
    ws[f"A{row_num+2}"] = "Date:-"
    ws[f"C{row_num+2}"] = "Signature:-"
    ws[f"A{row_num+2}"].font = ws[f"C{row_num+2}"].font = Font(underline='single')
    ws[f"C{row_num+2}"].font = ws[f"C{row_num+2}"].font = Font(underline='single')

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    # Save to BytesIO and send
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='attendance.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/stop_attendance', methods=['POST'])
def stop_attendance():
    global attendance_process
    if attendance_process and attendance_process.poll() is None:
        attendance_process.terminate()  # Sends SIGTERM
        attendance_process.wait(timeout=5)
        attendance_process = None
        return jsonify({'success': True, 'message': 'Attendance stopped'})
    else:
        return jsonify({'success': False, 'message': 'No running attendance process'})

@app.route('/get_lecture_dates')
def get_lecture_dates():
    course_id = request.args.get('course_id')
    unit_id = request.args.get('unit_id')
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    # Get the unitCode for the selected unit_id
    cursor.execute("SELECT unitCode FROM tblunit WHERE id = %s", (unit_id,))
    unit = cursor.fetchone()
    if not unit:
        cursor.close()
        connection.close()
        return jsonify([])
    unit_code = unit['unitCode']
    # Get the courseCode for the selected course_id
    cursor.execute("SELECT courseCode FROM tblcourse WHERE id = %s", (course_id,))
    course = cursor.fetchone()
    if not course:
        cursor.close()
        connection.close()
        return jsonify([])
    course_code = course['courseCode']
    # Now get all unique dates where attendance was launched for this course/unit
    cursor.execute("""
        SELECT DISTINCT DATE(dateMarked) as dateMarked
        FROM tblattendance
        WHERE course = %s AND unit = %s
        ORDER BY dateMarked DESC
    """, (course_code, unit_code))
    dates = [row['dateMarked'].strftime('%Y-%m-%d') for row in cursor.fetchall()]
    cursor.close()
    connection.close()
    return jsonify(dates)

@app.route('/export_view_attendance_excel')
def export_view_attendance_excel():
    course_id = request.args.get('course_id')
    unit_id = request.args.get('unit_id')
    date = request.args.get('date')

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    # Fetch course and unit names
    cursor.execute("SELECT name FROM tblcourse WHERE id = %s", (course_id,))
    course_row = cursor.fetchone()
    course_name = course_row['name'] if course_row else ''
    cursor.execute("SELECT name FROM tblunit WHERE id = %s", (unit_id,))
    unit_row = cursor.fetchone()
    unit_name = unit_row['name'] if unit_row else ''

    cursor.execute("""
        SELECT s.registrationNumber, CONCAT(s.firstName, ' ', s.lastName) as name, 
               IFNULL(a.attendanceStatus, 'Absent') as attendanceStatus
        FROM tblstudents s
        JOIN tblcourse c ON s.courseCode = c.courseCode
        JOIN tblunit u ON u.courseID = c.courseCode
        LEFT JOIN tblattendance a 
            ON a.studentRegistrationNumber = s.registrationNumber 
            AND a.unit = u.unitCode 
            AND DATE(a.dateMarked) = %s
        WHERE c.id = %s AND u.id = %s
        ORDER BY s.registrationNumber
    """, (date, course_id, unit_id))
    data = cursor.fetchall()
    cursor.close()
    connection.close()

    # Prepare Excel workbook
    import openpyxl
    from openpyxl.styles import Alignment, Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # University name at the top
    ws.merge_cells('A1:C1')
    ws['A1'] = "South Eastern University of Sri Lanka"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Course, Unit, Date info (use names)
    ws['A4'] = f"Course: {course_name}"
    ws['B4'] = f"Unit: {unit_name}"
    ws['C4'] = f"Date: {date}"
    ws['A4'].font = ws['B4'].font = ws['C4'].font = Font(bold=True)

    # Table headers
    ws['A6'] = "Registration No"
    ws['B6'] = "Name"
    ws['C6'] = "Status"
    for col in ['A', 'B', 'C']:
        ws[f"{col}6"].font = Font(bold=True)
        ws[f"{col}6"].alignment = Alignment(horizontal='center')

    # Attendance data
    row_num = 7
    for row in data:
        ws[f"A{row_num}"] = row['registrationNumber']
        ws[f"B{row_num}"] = row['name']
        ws[f"C{row_num}"] = row['attendanceStatus']
        row_num += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='view_attendance.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export_download_attendance_excel')
def export_download_attendance_excel():
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from io import BytesIO

    course_id = request.args.get('course_id')
    unit_id = request.args.get('unit_id')

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    # Get course, unit, and faculty names
    cursor.execute("SELECT name, facultyID FROM tblcourse WHERE id = %s", (course_id,))
    course_row = cursor.fetchone()
    course_name = course_row['name'] if course_row else ''
    faculty_code = course_row['facultyID'] if course_row else ''
    faculty_name = ''
    if faculty_code:
        cursor.execute("SELECT facultyName FROM tblfaculty WHERE facultyCode = %s", (faculty_code,))
        faculty_row = cursor.fetchone()
        faculty_name = faculty_row['facultyName'] if faculty_row else ''
    cursor.execute("SELECT name FROM tblunit WHERE id = %s", (unit_id,))
    unit_row = cursor.fetchone()
    unit_name = unit_row['name'] if unit_row else ''

    # Get all unique dates for this course/unit
    cursor.execute("""
        SELECT DISTINCT DATE(ta.dateMarked) as dateMarked
        FROM tblattendance ta
        JOIN tblunit u ON ta.unit = u.unitCode
        JOIN tblcourse c ON u.courseID = c.courseCode
        WHERE c.id = %s AND u.id = %s
        ORDER BY dateMarked
    """, (course_id, unit_id))
    dates = [row['dateMarked'].strftime('%Y-%m-%d') for row in cursor.fetchall()]

    # Get all students for this course/unit, including names
    cursor.execute("""
        SELECT s.registrationNumber, CONCAT(s.firstName, ' ', s.lastName) as name
        FROM tblstudents s
        JOIN tblcourse c ON s.courseCode = c.courseCode
        JOIN tblunit u ON u.courseID = c.courseCode
        WHERE c.id = %s AND u.id = %s
    """, (course_id, unit_id))
    students = cursor.fetchall()

    # Get all attendance records for this course/unit
    cursor.execute("""
        SELECT ta.studentRegistrationNumber, DATE(ta.dateMarked) as dateMarked, ta.attendanceStatus
        FROM tblattendance ta
        JOIN tblunit u ON ta.unit = u.unitCode
        JOIN tblcourse c ON u.courseID = c.courseCode
        WHERE c.id = %s AND u.id = %s
    """, (course_id, unit_id))
    records = cursor.fetchall()
    cursor.close()
    connection.close()

    # Build matrix
    student_map = {reg['registrationNumber']: {"registrationNumber": reg['registrationNumber'], "name": reg['name'], "attendance": {}} for reg in students}
    for rec in records:
        reg = rec['studentRegistrationNumber']
        date = rec['dateMarked'].strftime('%Y-%m-%d') if hasattr(rec['dateMarked'], 'strftime') else rec['dateMarked']
        status = rec['attendanceStatus']
        if reg in student_map:
            student_map[reg]["attendance"][date] = status

    # Fill in "Absent" for missing dates and calculate percentage
    for student in student_map.values():
        present_count = 0
        for date in dates:
            if date not in student["attendance"]:
                student["attendance"][date] = "Absent"
            if student["attendance"][date] == "Present":
                present_count += 1
        total = len(dates)
        student["percentage"] = f"{(present_count / total * 100):.2f}%" if total > 0 else "N/A"

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Download Attendance"

    # University name at the very top (merged and bold)
    col_count = 2 + len(dates) + 1  # Registration No, Name, dates, Percentage
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws.cell(row=1, column=1).value = "South Eastern University of Sri Lanka"
    ws.cell(row=1, column=1).font = Font(size=16, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # Faculty name (merged and bold)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws.cell(row=2, column=1).value = f"{faculty_name}"
    ws.cell(row=2, column=1).font = Font(size=13, bold=True)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # Course and unit name (merged and bold)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=col_count)
    ws.cell(row=4, column=1).value = f"Course: {course_name}"
    ws.cell(row=4, column=1).font = Font(size=12, bold=True)
    ws.cell(row=4, column=1).alignment = Alignment(horizontal='left', vertical='center')
    
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=col_count)
    ws.cell(row=5, column=1).value = f"Unit: {unit_name}"
    ws.cell(row=5, column=1).font = Font(size=12, bold=True)
    ws.cell(row=5, column=1).alignment = Alignment(horizontal='left', vertical='center')
    # Header (row 7)
    ws.append([])  # Row 6 blank
    ws.append(["Registration No", "Name"] + dates + ["Percentage of Attendance"])
    for col in range(1, col_count + 1):
        ws.cell(row=7, column=col).font = Font(bold=True)
        ws.cell(row=7, column=col).alignment = Alignment(horizontal='center')

    # Data rows (start from row 5)
    for student in student_map.values():
        row = [student["registrationNumber"], student["name"]]
        row += [student["attendance"][date] for date in dates]
        row.append(student["percentage"])
        ws.append(row)

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    for i in range(3, 3 + len(dates)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 15
    ws.column_dimensions[openpyxl.utils.get_column_letter(3 + len(dates))].width = 25

    # Save to BytesIO and send
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='download_attendance.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(debug=True)
